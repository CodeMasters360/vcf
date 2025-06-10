import sys
import binascii
from base64 import b64decode
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTreeWidget, QTreeWidgetItem, QLineEdit, QPushButton, QLabel,
    QFileDialog, QMessageBox, QMenu, QMenuBar, QStatusBar, QScrollArea,
    QHeaderView, QTabWidget, QSplitter, QTextEdit, QComboBox, QCheckBox,
    QGroupBox
)
from PyQt6.QtGui import QAction, QPixmap, QImage, QGuiApplication, QBrush, QColor, QFont
from PyQt6.QtCore import Qt, QSize
from PIL import Image
import io

# Excel export functionality
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class Contact:
    def __init__(self, data):
        self.name = data['name']
        self.phone = data['phone']
        self.additional_phones = data['additional_phones']
        self.original_lines = data['original_lines']
        self.has_photo = data['has_photo']
        self.photo_data = data['photo_data']
        self.selected = False

class VcfParser:
    def parse_vcf(self, vcf_content):
        lines = vcf_content.strip().split('\n')
        vcard_entries = []
        current_entry = []
        in_vcard = False

        for line in lines:
            line = line.strip()
            if line.startswith('BEGIN:VCARD'):
                if in_vcard:
                    vcard_entries.append(current_entry)
                current_entry = [line]
                in_vcard = True
            elif line.startswith('END:VCARD'):
                current_entry.append(line)
                vcard_entries.append(current_entry)
                current_entry = []
                in_vcard = False
            elif in_vcard:
                current_entry.append(line)

        if in_vcard and current_entry:
            vcard_entries.append(current_entry)

        contacts = []
        for entry in vcard_entries:
            processed_lines = []
            current_line = None

            for line in entry:
                if line.startswith('='):
                    if current_line is not None:
                        current_line += line[1:]
                else:
                    if current_line is not None:
                        processed_lines.append(current_line)
                    current_line = line

            if current_line is not None:
                processed_lines.append(current_line)

            name = None
            fn_name = None  # Full Name from FN field
            phones = []
            photo_data = None
            original_lines = entry.copy()
            has_photo = False

            for idx, line in enumerate(processed_lines):
                if line.startswith('END:VCARD'):
                    break
                if ':' not in line:
                    continue
                
                # Split only on the first colon to handle values with colons
                colon_index = line.find(':')
                if colon_index == -1:
                    continue
                    
                key = line[:colon_index]
                value = line[colon_index + 1:]
                key = key.upper()

                # Handle N field (structured name)
                if key.startswith('N'):
                    if 'CHARSET=UTF-8' in key and 'ENCODING=QUOTED-PRINTABLE' in key:
                        value = value.replace('==', '=')
                        try:
                            decoded_bytes = binascii.a2b_qp(value)
                            name = decoded_bytes.decode('utf-8').replace(';', ' ')
                        except Exception as e:
                            name = f"Error decoding N: {e}"
                    else:
                        # Handle simple N field format (e.g., N:;gffk;;;)
                        # N field format: Family;Given;Additional;Prefix;Suffix
                        name_parts = value.split(';')
                        name_components = []
                        
                        # Extract non-empty parts
                        for i, part in enumerate(name_parts[:5]):  # Only take first 5 parts
                            if part.strip():
                                name_components.append(part.strip())
                        
                        if name_components:
                            name = ' '.join(name_components)
                        else:
                            # If N field is empty or only semicolons, we'll use FN later
                            name = None
                
                # Handle FN field (formatted/full name)
                elif key.startswith('FN'):
                    if 'CHARSET=UTF-8' in key and 'ENCODING=QUOTED-PRINTABLE' in key:
                        value = value.replace('==', '=')
                        try:
                            decoded_bytes = binascii.a2b_qp(value)
                            fn_name = decoded_bytes.decode('utf-8')
                        except Exception as e:
                            fn_name = f"Error decoding FN: {e}"
                    else:
                        fn_name = value.strip()
                
                # Handle telephone numbers
                elif key.startswith('TEL'):
                    phones.append(value)
                
                # Handle photos
                elif key.startswith('PHOTO'):
                    has_photo = True
                    if 'BASE64' in key:
                        photo_lines = [value]
                        next_idx = idx + 1
                        while next_idx < len(processed_lines):
                            next_line = processed_lines[next_idx]
                            if next_line.startswith(' ') or ':' not in next_line:
                                photo_lines.append(next_line.strip())
                                next_idx += 1
                            else:
                                break
                        photo_data = ''.join(photo_lines).replace(' ', '').replace('\n', '')

            # Determine the final name to use
            final_name = None
            
            # Priority: 1. Parsed N field, 2. FN field, 3. Skip if both empty
            if name and name.strip():
                final_name = name.strip()
            elif fn_name and fn_name.strip():
                final_name = fn_name.strip()
            
            # Only create contact if we have a name
            if final_name:
                main_phone = phones[0] if phones else None
                additional_phones = ', '.join(phones[1:]) if len(phones) > 1 else ''
                contacts.append(Contact({
                    'name': final_name.replace('ي', 'ی').replace('ك', 'ک'),
                    'phone': main_phone,
                    'additional_phones': additional_phones,
                    'original_lines': original_lines,
                    'has_photo': has_photo,
                    'photo_data': photo_data
                }))

        return contacts

class ExcelExporter:
    """Class to handle Excel export functionality"""
    
    @staticmethod
    def export_contacts_to_excel(contacts, file_path, sheet_name="Contacts", include_metadata=True):
        """Export contacts to Excel file"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl library is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Define headers
        headers = ['#', 'Name', 'Phone', 'Additional Phones', 'Has Photo']
        
        # Style the headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write contact data
        for row, contact in enumerate(contacts, 2):
            ws.cell(row=row, column=1, value=row-1)  # Index
            ws.cell(row=row, column=2, value=contact.name)
            ws.cell(row=row, column=3, value=contact.phone or "No Phone")
            ws.cell(row=row, column=4, value=contact.additional_phones or "-")
            ws.cell(row=row, column=5, value="Yes" if contact.has_photo else "No")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(contacts) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            # Set column width with some padding
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Add metadata if requested
        if include_metadata:
            metadata_row = len(contacts) + 3
            ws.cell(row=metadata_row, column=1, value="Export Information:")
            ws.cell(row=metadata_row, column=1).font = Font(bold=True)
            
            ws.cell(row=metadata_row + 1, column=1, value=f"Total Contacts: {len(contacts)}")
            ws.cell(row=metadata_row + 2, column=1, value=f"Export Date: 2025-06-10 01:00:27 UTC")
            ws.cell(row=metadata_row + 3, column=1, value="Exported by: VCF Viewer Tool")
            ws.cell(row=metadata_row + 4, column=1, value="Created by: CodeMasters360")
        
        # Save the workbook
        wb.save(file_path)
    
    @staticmethod
    def export_comparison_to_excel(comparison_results, file_path, match_method, phone_filter):
        """Export comparison results to Excel with multiple sheets"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl library is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        ExcelExporter._create_summary_sheet(summary_ws, comparison_results, match_method, phone_filter)
        
        # Create sheets for each category
        if comparison_results['only_in_file1']:
            file1_ws = wb.create_sheet("Only in File 1")
            ExcelExporter._create_contacts_sheet(file1_ws, comparison_results['only_in_file1'], "Only in File 1")
        
        if comparison_results['only_in_file2']:
            file2_ws = wb.create_sheet("Only in File 2")
            ExcelExporter._create_contacts_sheet(file2_ws, comparison_results['only_in_file2'], "Only in File 2")
        
        if comparison_results['common']:
            common_ws = wb.create_sheet("Common Contacts")
            ExcelExporter._create_common_contacts_sheet(common_ws, comparison_results['common'])
        
        wb.save(file_path)
    
    @staticmethod
    def _create_summary_sheet(ws, results, match_method, phone_filter):
        """Create summary sheet for comparison results"""
        # Title
        ws.cell(row=1, column=1, value="VCF File Comparison Summary")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        # Comparison details
        ws.cell(row=3, column=1, value="Comparison Details:")
        ws.cell(row=3, column=1).font = Font(bold=True)
        ws.cell(row=4, column=1, value=f"Match Method: {match_method}")
        ws.cell(row=5, column=1, value=f"Phone Filter: {phone_filter}")
        
        # File statistics
        ws.cell(row=7, column=1, value="File Statistics:")
        ws.cell(row=7, column=1).font = Font(bold=True)
        ws.cell(row=8, column=1, value=f"File 1 Total Contacts: {results['file1_total']}")
        ws.cell(row=9, column=1, value=f"File 1 Filtered Contacts: {results['file1_filtered']}")
        ws.cell(row=10, column=1, value=f"File 2 Total Contacts: {results['file2_total']}")
        ws.cell(row=11, column=1, value=f"File 2 Filtered Contacts: {results['file2_filtered']}")
        
        # Results
        ws.cell(row=13, column=1, value="Comparison Results:")
        ws.cell(row=13, column=1).font = Font(bold=True)
        ws.cell(row=14, column=1, value=f"Contacts only in File 1: {len(results['only_in_file1'])}")
        ws.cell(row=15, column=1, value=f"Contacts only in File 2: {len(results['only_in_file2'])}")
        ws.cell(row=16, column=1, value=f"Common contacts: {len(results['common'])}")
        
        # Export info
        ws.cell(row=18, column=1, value="Export Information:")
        ws.cell(row=18, column=1).font = Font(bold=True)
        ws.cell(row=19, column=1, value=f"Export Date: 2025-06-10 01:00:27 UTC")
        ws.cell(row=20, column=1, value="Exported by: VCF Viewer Tool")
        ws.cell(row=21, column=1, value="Created by: CodeMasters360")
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 40
    
    @staticmethod
    def _create_contacts_sheet(ws, contacts, sheet_title):
        """Create a sheet for single contacts list"""
        headers = ['#', 'Name', 'Phone', 'Additional Phones', 'Has Photo']
        
        # Style the headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write contact data
        for row, contact in enumerate(contacts, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=contact.name)
            ws.cell(row=row, column=3, value=contact.phone or "No Phone")
            ws.cell(row=row, column=4, value=contact.additional_phones or "-")
            ws.cell(row=row, column=5, value="Yes" if contact.has_photo else "No")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(contacts) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @staticmethod
    def _create_common_contacts_sheet(ws, common_contacts):
        """Create a sheet for common contacts"""
        headers = ['#', 'Name (File 1)', 'Phone (File 1)', 'Additional Phones (File 1)', 
                  'Name (File 2)', 'Phone (File 2)', 'Additional Phones (File 2)']
        
        # Style the headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write contact data
        for row, (contact1, contact2) in enumerate(common_contacts, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=contact1.name)
            ws.cell(row=row, column=3, value=contact1.phone or "No Phone")
            ws.cell(row=row, column=4, value=contact1.additional_phones or "-")
            ws.cell(row=row, column=5, value=contact2.name)
            ws.cell(row=row, column=6, value=contact2.phone or "No Phone")
            ws.cell(row=row, column=7, value=contact2.additional_phones or "-")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(common_contacts) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

class VcfComparator:
    def __init__(self):
        self.file1_path = ""
        self.file2_path = ""
        self.file1_contacts = []
        self.file2_contacts = []
        
    def normalize_phone(self, phone):
        """Normalize phone number for comparison"""
        if not phone:
            return ""
        # Remove common phone number separators and spaces
        normalized = ''.join(c for c in phone if c.isdigit() or c == '+')
        # Remove leading zeros and country codes for better matching
        if normalized.startswith('+'):
            return normalized
        return normalized.lstrip('0')
    
    def normalize_name(self, name):
        """Normalize name for comparison"""
        if not name:
            return ""
        return name.lower().strip().replace('ي', 'ی').replace('ك', 'ک')
    
    def contacts_match(self, contact1, contact2, match_method):
        """Check if two contacts match based on the selected method"""
        if match_method == "Name + Phone":
            return (self.normalize_name(contact1.name) == self.normalize_name(contact2.name) and 
                    self.normalize_phone(contact1.phone) == self.normalize_phone(contact2.phone))
        elif match_method == "Name Only":
            return self.normalize_name(contact1.name) == self.normalize_name(contact2.name)
        elif match_method == "Phone Only":
            return self.normalize_phone(contact1.phone) == self.normalize_phone(contact2.phone)
        return False
    
    def find_contact_in_list(self, target_contact, contact_list, match_method):
        """Find if a contact exists in a list using the specified matching method"""
        for contact in contact_list:
            if self.contacts_match(target_contact, contact, match_method):
                return contact
        return None
    
    def filter_contacts_by_phone(self, contacts, phone_filter):
        """Filter contacts based on phone number criteria"""
        if phone_filter == "All Contacts":
            return contacts
        elif phone_filter == "With Phone Only":
            return [c for c in contacts if c.phone and c.phone.strip()]
        elif phone_filter == "Without Phone Only":
            return [c for c in contacts if not c.phone or not c.phone.strip()]
        else:
            return contacts
    
    def compare_files(self, file1_contacts, file2_contacts, match_method="Name + Phone", phone_filter="All Contacts"):
        """Compare two lists of contacts and return differences"""
        # Apply phone filter to both contact lists before comparison
        filtered_file1_contacts = self.filter_contacts_by_phone(file1_contacts, phone_filter)
        filtered_file2_contacts = self.filter_contacts_by_phone(file2_contacts, phone_filter)
        
        self.file1_contacts = filtered_file1_contacts
        self.file2_contacts = filtered_file2_contacts
        
        # Find contacts only in file1
        only_in_file1 = []
        for contact in filtered_file1_contacts:
            if not self.find_contact_in_list(contact, filtered_file2_contacts, match_method):
                only_in_file1.append(contact)
        
        # Find contacts only in file2
        only_in_file2 = []
        for contact in filtered_file2_contacts:
            if not self.find_contact_in_list(contact, filtered_file1_contacts, match_method):
                only_in_file2.append(contact)
        
        # Find common contacts
        common_contacts = []
        for contact in filtered_file1_contacts:
            match = self.find_contact_in_list(contact, filtered_file2_contacts, match_method)
            if match:
                common_contacts.append((contact, match))
        
        return {
            'only_in_file1': only_in_file1,
            'only_in_file2': only_in_file2,
            'common': common_contacts,
            'file1_total': len(file1_contacts),
            'file2_total': len(file2_contacts),
            'file1_filtered': len(filtered_file1_contacts),
            'file2_filtered': len(filtered_file2_contacts),
            'phone_filter': phone_filter
        }

class SortableTreeWidget(QTreeWidget):
    """Custom TreeWidget with sorting functionality"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.sort_column = 0
        self.sort_order = Qt.SortOrder.AscendingOrder
        self.original_data = []
        self.current_data = []
        self.data_type = None  # 'single' for single contacts, 'tuple' for contact pairs
        self.setSortingEnabled(False)
        self.header().sectionClicked.connect(self.handle_header_click)
        self.header().setSectionsClickable(True)
        self.header().setSortIndicatorShown(True)

    def set_data(self, data, headers):
        """Set the data for the tree widget"""
        self.original_data = data.copy()
        self.current_data = data.copy()
        
        # Determine data type based on first element
        if data:
            if hasattr(data[0], 'name'):
                self.data_type = 'single'
            else:
                self.data_type = 'tuple'
        else:
            self.data_type = None  # Empty data
            
        self.setHeaderLabels(headers)
        self.display_data()

    def handle_header_click(self, logical_index):
        """Handle header click for sorting"""
        # Don't sort if there's no data
        if not self.current_data:
            return
            
        if logical_index == self.sort_column:
            self.sort_order = (
                Qt.SortOrder.DescendingOrder 
                if self.sort_order == Qt.SortOrder.AscendingOrder 
                else Qt.SortOrder.AscendingOrder
            )
        else:
            self.sort_column = logical_index
            self.sort_order = Qt.SortOrder.AscendingOrder

        self.sort_data()
        self.display_data()

    def sort_data(self):
        """Sort the data based on the selected column"""
        # Don't sort if there's no data
        if not self.current_data:
            return
            
        if self.sort_column == 0:  # Row number column - reset to original order
            self.current_data = self.original_data.copy()
        else:
            # Get the appropriate sort key based on column and data type
            sort_key = None
            
            if self.data_type == 'single':
                # Single contact data
                if self.sort_column == 1:  # Name column
                    sort_key = lambda x: x.name.lower()
                elif self.sort_column == 2:  # Phone column
                    sort_key = lambda x: x.phone or ''
                elif self.sort_column == 3:  # Additional phones column
                    sort_key = lambda x: len(x.additional_phones.split(', ')) if x.additional_phones else 0
            elif self.data_type == 'tuple':
                # Tuple of contacts (common contacts)
                if self.sort_column == 1:  # First contact name
                    sort_key = lambda x: x[0].name.lower()
                elif self.sort_column == 2:  # First contact phone
                    sort_key = lambda x: x[0].phone or ''
                elif self.sort_column == 3:  # Second contact name
                    sort_key = lambda x: x[1].name.lower()
                elif self.sort_column == 4:  # Second contact phone
                    sort_key = lambda x: x[1].phone or ''
            
            # Apply sorting if we have a valid sort key
            if sort_key:
                try:
                    self.current_data.sort(
                        key=sort_key, 
                        reverse=self.sort_order == Qt.SortOrder.DescendingOrder
                    )
                except Exception as e:
                    # If sorting fails, keep the current order
                    print(f"Sorting error: {e}")

    def display_data(self):
        """Display the current data in the tree widget"""
        self.clear()
        
        if not self.current_data:
            # Update header to show no sort indicator when empty
            self.header().setSortIndicator(-1, self.sort_order)
            return
        
        for index, item in enumerate(self.current_data, start=1):
            if self.data_type == 'single':
                # Single contact
                tree_item = QTreeWidgetItem([
                    str(index),
                    item.name,
                    item.phone or 'No Phone',
                    item.additional_phones or '-'
                ])
            elif self.data_type == 'tuple':
                # Tuple of contacts (common contacts)
                contact1, contact2 = item
                tree_item = QTreeWidgetItem([
                    str(index),
                    contact1.name,
                    contact1.phone or 'No Phone',
                    contact2.name,
                    contact2.phone or 'No Phone'
                ])
            else:
                continue  # Skip if data type is unknown
            
            # Store the original data for export functionality
            tree_item.setData(0, Qt.ItemDataRole.UserRole, item)
            self.addTopLevelItem(tree_item)
        
        # Update header sort indicator
        if self.sort_column != 0:
            self.header().setSortIndicator(self.sort_column, self.sort_order)
        else:
            self.header().setSortIndicator(-1, self.sort_order)

    def get_current_data(self):
        """Get the current sorted data"""
        return self.current_data

class ComparisonWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.comparator = VcfComparator()
        self.comparison_results = None
        self.initUI()
    
    def initUI(self):
        self.setWindowTitle('VCF File Comparison')
        self.setGeometry(150, 150, 1200, 850)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        # File selection section
        file_section = QHBoxLayout()
        
        # File 1 selection
        file1_layout = QVBoxLayout()
        file1_layout.addWidget(QLabel("File 1:"))
        self.file1_label = QLabel("No file selected")
        self.file1_label.setStyleSheet("border: 1px solid gray; padding: 5px;")
        self.file1_btn = QPushButton("Select File 1")
        self.file1_btn.clicked.connect(self.select_file1)
        file1_layout.addWidget(self.file1_label)
        file1_layout.addWidget(self.file1_btn)
        
        # File 2 selection
        file2_layout = QVBoxLayout()
        file2_layout.addWidget(QLabel("File 2:"))
        self.file2_label = QLabel("No file selected")
        self.file2_label.setStyleSheet("border: 1px solid gray; padding: 5px;")
        self.file2_btn = QPushButton("Select File 2")
        self.file2_btn.clicked.connect(self.select_file2)
        file2_layout.addWidget(self.file2_label)
        file2_layout.addWidget(self.file2_btn)
        
        file_section.addLayout(file1_layout)
        file_section.addLayout(file2_layout)
        
        # Comparison options section
        options_group = QGroupBox("Comparison Options")
        options_layout = QVBoxLayout(options_group)
        
        # First row: Match method
        match_layout = QHBoxLayout()
        match_layout.addWidget(QLabel("Match Method:"))
        self.match_method_combo = QComboBox()
        self.match_method_combo.addItems(["Name + Phone", "Name Only", "Phone Only"])
        match_layout.addWidget(self.match_method_combo)
        match_layout.addStretch()
        
        # Second row: Phone filter
        phone_layout = QHBoxLayout()
        phone_layout.addWidget(QLabel("Phone Filter:"))
        self.phone_filter_combo = QComboBox()
        self.phone_filter_combo.addItems(["All Contacts", "With Phone Only", "Without Phone Only"])
        self.phone_filter_combo.setToolTip("Choose which contacts to include in the comparison based on phone number presence")
        phone_layout.addWidget(self.phone_filter_combo)
        phone_layout.addStretch()
        
        # Third row: Compare button
        button_layout = QHBoxLayout()
        self.compare_btn = QPushButton("Compare Files")
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        button_layout.addWidget(self.compare_btn)
        button_layout.addStretch()
        
        options_layout.addLayout(match_layout)
        options_layout.addLayout(phone_layout)
        options_layout.addLayout(button_layout)
        
        layout.addLayout(file_section)
        layout.addWidget(options_group)
        
        # Results section
        self.results_tab = QTabWidget()
        
        # Summary tab
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.results_tab.addTab(self.summary_text, "Summary")
        
        # Only in File 1 tab
        self.file1_tree = SortableTreeWidget()
        self.file1_tree.setHeaderLabels(['#', 'Name', 'Phone', 'Additional Phones'])
        self.results_tab.addTab(self.file1_tree, "Only in File 1")
        
        # Only in File 2 tab
        self.file2_tree = SortableTreeWidget()
        self.file2_tree.setHeaderLabels(['#', 'Name', 'Phone', 'Additional Phones'])
        self.results_tab.addTab(self.file2_tree, "Only in File 2")
        
        # Common contacts tab
        self.common_tree = SortableTreeWidget()
        self.common_tree.setHeaderLabels(['#', 'Name (File 1)', 'Phone (File 1)', 'Name (File 2)', 'Phone (File 2)'])
        self.results_tab.addTab(self.common_tree, "Common Contacts")
        
        layout.addWidget(self.results_tab)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        # VCF Export buttons
        self.export_file1_btn = QPushButton("Export VCF: Only in File 1")
        self.export_file1_btn.clicked.connect(lambda: self.export_contacts('file1', 'vcf'))
        self.export_file1_btn.setEnabled(False)
        
        self.export_file2_btn = QPushButton("Export VCF: Only in File 2")
        self.export_file2_btn.clicked.connect(lambda: self.export_contacts('file2', 'vcf'))
        self.export_file2_btn.setEnabled(False)
        
        self.export_common_btn = QPushButton("Export VCF: Common Contacts")
        self.export_common_btn.clicked.connect(lambda: self.export_contacts('common', 'vcf'))
        self.export_common_btn.setEnabled(False)
        
        export_layout.addWidget(self.export_file1_btn)
        export_layout.addWidget(self.export_file2_btn)
        export_layout.addWidget(self.export_common_btn)
        
        # Excel Export buttons
        excel_layout = QHBoxLayout()
        
        self.export_excel_file1_btn = QPushButton("Export Excel: Only in File 1")
        self.export_excel_file1_btn.clicked.connect(lambda: self.export_contacts('file1', 'excel'))
        self.export_excel_file1_btn.setEnabled(False)
        
        self.export_excel_file2_btn = QPushButton("Export Excel: Only in File 2")
        self.export_excel_file2_btn.clicked.connect(lambda: self.export_contacts('file2', 'excel'))
        self.export_excel_file2_btn.setEnabled(False)
        
        self.export_excel_common_btn = QPushButton("Export Excel: Common Contacts")
        self.export_excel_common_btn.clicked.connect(lambda: self.export_contacts('common', 'excel'))
        self.export_excel_common_btn.setEnabled(False)
        
        self.export_excel_all_btn = QPushButton("Export Excel: Complete Comparison")
        self.export_excel_all_btn.clicked.connect(lambda: self.export_contacts('all', 'excel'))
        self.export_excel_all_btn.setEnabled(False)
        
        excel_layout.addWidget(self.export_excel_file1_btn)
        excel_layout.addWidget(self.export_excel_file2_btn)
        excel_layout.addWidget(self.export_excel_common_btn)
        excel_layout.addWidget(self.export_excel_all_btn)
        
        excel_layout.addStretch()
        export_layout.addStretch()
        
        layout.addLayout(export_layout)
        layout.addLayout(excel_layout)
        
        # Show Excel availability status
        if not EXCEL_AVAILABLE:
            excel_warning = QLabel("⚠️ Excel export requires 'openpyxl' library. Install with: pip install openpyxl")
            excel_warning.setStyleSheet("color: orange; font-weight: bold;")
            layout.addWidget(excel_warning)
            
            # Disable Excel buttons
            self.export_excel_file1_btn.setEnabled(False)
            self.export_excel_file2_btn.setEnabled(False)
            self.export_excel_common_btn.setEnabled(False)
            self.export_excel_all_btn.setEnabled(False)
    
    def select_file1(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select VCF File 1", "", "VCF Files (*.vcf)")
        if file_path:
            self.comparator.file1_path = file_path
            self.file1_label.setText(file_path.split('/')[-1])
            self.check_ready_to_compare()
    
    def select_file2(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select VCF File 2", "", "VCF Files (*.vcf)")
        if file_path:
            self.comparator.file2_path = file_path
            self.file2_label.setText(file_path.split('/')[-1])
            self.check_ready_to_compare()
    
    def check_ready_to_compare(self):
        if self.comparator.file1_path and self.comparator.file2_path:
            self.compare_btn.setEnabled(True)
    
    def compare_files(self):
        try:
            # Parse file 1
            with open(self.comparator.file1_path, 'r', encoding='utf-8') as f:
                content1 = f.read()
            parser = VcfParser()
            file1_contacts = parser.parse_vcf(content1)
            
            # Parse file 2
            with open(self.comparator.file2_path, 'r', encoding='utf-8') as f:
                content2 = f.read()
            file2_contacts = parser.parse_vcf(content2)
            
            # Compare files
            match_method = self.match_method_combo.currentText()
            phone_filter = self.phone_filter_combo.currentText()
            self.comparison_results = self.comparator.compare_files(
                file1_contacts, file2_contacts, match_method, phone_filter
            )
            
            # Display results
            self.display_results()
            
            # Enable export buttons
            self.export_file1_btn.setEnabled(True)
            self.export_file2_btn.setEnabled(True)
            self.export_common_btn.setEnabled(True)
            
            if EXCEL_AVAILABLE:
                self.export_excel_file1_btn.setEnabled(True)
                self.export_excel_file2_btn.setEnabled(True)
                self.export_excel_common_btn.setEnabled(True)
                self.export_excel_all_btn.setEnabled(True)
            
        except Exception as e:
            QMessageBox.critical(self, "Comparison Error", f"Error comparing files: {str(e)}")
    
    def display_results(self):
        if not self.comparison_results:
            return
        
        results = self.comparison_results
        
        # Generate filter description
        phone_filter_desc = ""
        if results['phone_filter'] == "With Phone Only":
            phone_filter_desc = " (contacts with phone numbers only)"
        elif results['phone_filter'] == "Without Phone Only":
            phone_filter_desc = " (contacts without phone numbers only)"
        elif results['phone_filter'] == "All Contacts":
            phone_filter_desc = " (all contacts)"
        
        # Update summary
        summary = f"""
Comparison Results
==================

File 1: {self.comparator.file1_path.split('/')[-1]}
Total contacts: {results['file1_total']}
Filtered contacts: {results['file1_filtered']}{phone_filter_desc}

File 2: {self.comparator.file2_path.split('/')[-1]}
Total contacts: {results['file2_total']}
Filtered contacts: {results['file2_filtered']}{phone_filter_desc}

Match Method: {self.match_method_combo.currentText()}
Phone Filter: {results['phone_filter']}

Results{phone_filter_desc}:
--------
Contacts only in File 1: {len(results['only_in_file1'])}
Contacts only in File 2: {len(results['only_in_file2'])}
Common contacts: {len(results['common'])}

Analysis:
---------
File 1 has {len(results['only_in_file1'])} unique contacts not found in File 2
File 2 has {len(results['only_in_file2'])} unique contacts not found in File 1
{len(results['common'])} contacts are present in both files

Filter Statistics:
------------------
File 1: {results['file1_total'] - results['file1_filtered']} contacts excluded by phone filter
File 2: {results['file2_total'] - results['file2_filtered']} contacts excluded by phone filter

Export Options:
---------------
• VCF Export: Maintains original contact format, compatible with phone/email apps
• Excel Export: Creates structured spreadsheet with enhanced formatting and metadata
• Complete Comparison Excel: All results in one file with multiple sheets

Sorting Instructions:
--------------------
- Click any column header to sort by that column
- Click the same header again to reverse the sort order
- Click the # column to restore original order
- All tables support sorting by Name and Phone number

VCF Format Support:
-------------------
- VCF 2.1 and 3.0 formats supported
- Handles both N (structured name) and FN (full name) fields
- Supports UTF-8 and Quoted-Printable encoding
- Fallback from N field to FN field for name extraction
        """
        
        self.summary_text.setPlainText(summary)
        
        # Update trees with sortable data
        self.file1_tree.set_data(results['only_in_file1'], ['#', 'Name', 'Phone', 'Additional Phones'])
        self.file2_tree.set_data(results['only_in_file2'], ['#', 'Name', 'Phone', 'Additional Phones'])
        self.common_tree.set_data(results['common'], ['#', 'Name (File 1)', 'Phone (File 1)', 'Name (File 2)', 'Phone (File 2)'])
        
        # Update tab titles with counts and filter info
        filter_suffix = ""
        if results['phone_filter'] != "All Contacts":
            if results['phone_filter'] == "With Phone Only":
                filter_suffix = " (📞)"
            elif results['phone_filter'] == "Without Phone Only":
                filter_suffix = " (📵)"
        
        self.results_tab.setTabText(1, f"Only in File 1 ({len(results['only_in_file1'])}){filter_suffix}")
        self.results_tab.setTabText(2, f"Only in File 2 ({len(results['only_in_file2'])}){filter_suffix}")
        self.results_tab.setTabText(3, f"Common Contacts ({len(results['common'])}){filter_suffix}")
    
    def export_contacts(self, contact_type, export_format):
        if not self.comparison_results:
            return
        
        if export_format == 'excel':
            self.export_to_excel(contact_type)
        else:
            self.export_to_vcf(contact_type)
    
    def export_to_excel(self, contact_type):
        """Export contacts to Excel format"""
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(self, "Excel Export Error", 
                               "Excel export requires 'openpyxl' library.\n"
                               "Install it with: pip install openpyxl")
            return
        
        if contact_type == 'all':
            # Export complete comparison to Excel
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Complete Comparison to Excel", 
                "vcf_comparison_complete.xlsx", 
                "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
            
            try:
                ExcelExporter.export_comparison_to_excel(
                    self.comparison_results, 
                    file_path, 
                    self.match_method_combo.currentText(), 
                    self.phone_filter_combo.currentText()
                )
                QMessageBox.information(self, "Export Success", 
                                      f"Complete comparison exported to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")
            return
        
        # Single category export
        contacts_to_export = []
        default_filename = ""
        
        # Add filter suffix to filename
        filter_suffix = ""
        if self.comparison_results['phone_filter'] == "With Phone Only":
            filter_suffix = "_with_phone"
        elif self.comparison_results['phone_filter'] == "Without Phone Only":
            filter_suffix = "_without_phone"
        
        if contact_type == 'file1':
            current_data = self.file1_tree.get_current_data()
            contacts_to_export = current_data
            default_filename = f"only_in_file1{filter_suffix}.xlsx"
        elif contact_type == 'file2':
            current_data = self.file2_tree.get_current_data()
            contacts_to_export = current_data
            default_filename = f"only_in_file2{filter_suffix}.xlsx"
        elif contact_type == 'common':
            current_data = self.common_tree.get_current_data()
            contacts_to_export = [pair[0] for pair in current_data]
            default_filename = f"common_contacts{filter_suffix}.xlsx"
        
        if not contacts_to_export:
            QMessageBox.information(self, "Export", "No contacts to export.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Contacts to Excel", default_filename, "Excel Files (*.xlsx)")
        if not file_path:
            return
        
        try:
            ExcelExporter.export_contacts_to_excel(contacts_to_export, file_path)
            filter_msg = f" (filtered: {self.comparison_results['phone_filter']})" if self.comparison_results['phone_filter'] != "All Contacts" else ""
            QMessageBox.information(self, "Export Success", f"Exported {len(contacts_to_export)} contacts to {file_path}{filter_msg}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")
    
    def export_to_vcf(self, contact_type):
        """Export contacts to VCF format"""
        contacts_to_export = []
        default_filename = ""
        
        # Add filter suffix to filename
        filter_suffix = ""
        if self.comparison_results['phone_filter'] == "With Phone Only":
            filter_suffix = "_with_phone"
        elif self.comparison_results['phone_filter'] == "Without Phone Only":
            filter_suffix = "_without_phone"
        
        if contact_type == 'file1':
            # Get current sorted data from the tree widget
            current_data = self.file1_tree.get_current_data()
            contacts_to_export = current_data
            default_filename = f"only_in_file1{filter_suffix}.vcf"
        elif contact_type == 'file2':
            # Get current sorted data from the tree widget
            current_data = self.file2_tree.get_current_data()
            contacts_to_export = current_data
            default_filename = f"only_in_file2{filter_suffix}.vcf"
        elif contact_type == 'common':
            # For common contacts, export from file1 (first contact in each pair)
            current_data = self.common_tree.get_current_data()
            contacts_to_export = [pair[0] for pair in current_data]
            default_filename = f"common_contacts{filter_suffix}.vcf"
        
        if not contacts_to_export:
            QMessageBox.information(self, "Export", "No contacts to export.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Contacts", default_filename, "VCF Files (*.vcf)")
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                for contact in contacts_to_export:
                    in_photo = False
                    for line in contact.original_lines:
                        stripped_line = line.strip()
                        
                        if stripped_line == '':
                            f.write('\n')
                            continue
                            
                        if stripped_line.upper().startswith('PHOTO'):
                            f.write(stripped_line + '\n')
                            in_photo = True
                        elif in_photo:
                            if stripped_line.startswith('END:VCARD'):
                                f.write(line + '\n')
                                in_photo = False
                            else:
                                f.write(' ' + line.lstrip() + '\n')
                        else:
                            f.write(line + '\n')
            
            filter_msg = f" (filtered: {self.comparison_results['phone_filter']})" if self.comparison_results['phone_filter'] != "All Contacts" else ""
            QMessageBox.information(self, "Export Success", f"Exported {len(contacts_to_export)} contacts to {file_path}{filter_msg}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting contacts: {str(e)}")

class ContactViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.all_contacts = []
        self.contacts = []
        self.sort_column = 1
        self.sort_order = Qt.SortOrder.AscendingOrder
        self.comparison_window = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle('VCF Viewer')
        self.setGeometry(100, 100, 1000, 700)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(['#', 'Name', 'Phone', 'Additional Phones', 'Photo', 'Select'])
        self.tree.setSortingEnabled(False)  # Disable built-in sorting
        self.tree.header().sectionClicked.connect(self.handle_header_click)
        self.tree.itemDoubleClicked.connect(self.show_photo)
        self.tree.itemChanged.connect(self.handle_item_changed)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self.show_context_menu)

        header = self.tree.header()
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)

        self.tree.setColumnWidth(0, 50)
        self.tree.setColumnWidth(1, 200)
        self.tree.setColumnWidth(2, 150)
        self.tree.setColumnWidth(3, 120)  # Reduced from 200 to 120
        self.tree.setColumnWidth(4, 60)
        self.tree.setColumnWidth(5, 80)

        self.search_box = QLineEdit()
        self.search_box.textChanged.connect(self.filter_contacts)
        self.clear_btn = QPushButton('Clear')
        self.clear_btn.clicked.connect(self.clear_search)

        # Add selection control buttons
        self.select_all_btn = QPushButton('Select All')
        self.select_all_btn.clicked.connect(self.select_all)
        self.deselect_all_btn = QPushButton('Deselect All')
        self.deselect_all_btn.clicked.connect(self.deselect_all)
        self.invert_selection_btn = QPushButton('Invert Selection')
        self.invert_selection_btn.clicked.connect(self.invert_selection)

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setFixedSize(400, 400)  # Change from (300, 300) to desired size

        search_layout = QHBoxLayout()
        search_layout.addWidget(self.search_box)
        search_layout.addWidget(self.clear_btn)

        # Add selection buttons layout
        selection_layout = QHBoxLayout()
        selection_layout.addWidget(self.select_all_btn)
        selection_layout.addWidget(self.deselect_all_btn)
        selection_layout.addWidget(self.invert_selection_btn)
        selection_layout.addStretch()  # Push buttons to the left

        main_layout = QVBoxLayout()
        main_layout.addLayout(search_layout)
        main_layout.addLayout(selection_layout)
        main_layout.addWidget(self.tree)

        image_scroll = QScrollArea()
        image_scroll.setWidget(self.image_label)
        image_scroll.setWidgetResizable(True)

        central_widget = QWidget()
        central_layout = QHBoxLayout()
        central_layout.addLayout(main_layout, 70)
        central_layout.addWidget(image_scroll, 30)
        central_widget.setLayout(central_layout)
        self.setCentralWidget(central_widget)

        menubar = self.menuBar()
        file_menu = menubar.addMenu('File')

        import_action = QAction('Import VCF', self)
        import_action.triggered.connect(self.import_vcf)
        file_menu.addAction(import_action)

        save_action = QAction('Save VCF', self)
        save_action.triggered.connect(self.save_vcf)
        file_menu.addAction(save_action)
        
        # Add Excel export to main viewer
        if EXCEL_AVAILABLE:
            export_excel_action = QAction('Export to Excel', self)
            export_excel_action.triggered.connect(self.export_to_excel)
            file_menu.addAction(export_excel_action)

        delete_action = QAction('Delete Selected', self)
        delete_action.triggered.connect(self.delete_selected)
        file_menu.addAction(delete_action)

        delete_no_phone = QAction('Delete Without Phone', self)
        delete_no_phone.triggered.connect(self.delete_contacts_without_phone)
        file_menu.addAction(delete_no_phone)

        # Add comparison menu
        tools_menu = menubar.addMenu('Tools')
        compare_action = QAction('Compare VCF Files', self)
        compare_action.triggered.connect(self.open_comparison_window)
        tools_menu.addAction(compare_action)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # Add permanent widgets to status bar
        self.contact_count_label = QLabel("Contacts: 0")
        self.selected_count_label = QLabel("Selected: 0")
        self.status_bar.addPermanentWidget(self.contact_count_label)
        self.status_bar.addPermanentWidget(self.selected_count_label)
        
        # Show Excel status in status bar
        if not EXCEL_AVAILABLE:
            excel_status = QLabel("Excel export unavailable (install openpyxl)")
            excel_status.setStyleSheet("color: orange;")
            self.status_bar.addPermanentWidget(excel_status)

    def export_to_excel(self):
        """Export current contacts to Excel"""
        if not self.contacts:
            self.show_warning("Empty List", "No contacts to export")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "contacts.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            ExcelExporter.export_contacts_to_excel(self.contacts, file_path)
            self.status_bar.showMessage(f"Exported {len(self.contacts)} contacts to Excel")
            QMessageBox.information(self, "Export Success", f"Exported {len(self.contacts)} contacts to {file_path}")
        except Exception as e:
            self.show_error("Excel Export Error", str(e))

    def open_comparison_window(self):
        if self.comparison_window is None:
            self.comparison_window = ComparisonWindow(self)
        self.comparison_window.show()
        self.comparison_window.raise_()
        self.comparison_window.activateWindow()

    def handle_header_click(self, logical_index):
        self.sort_contacts(logical_index)

    def import_vcf(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open VCF", "", "VCF Files (*.vcf)")
        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            parser = VcfParser()
            self.all_contacts = parser.parse_vcf(content)
            self.contacts = self.all_contacts.copy()
            self.display_contacts()
            self.status_bar.showMessage("File loaded successfully")
        except Exception as e:
            self.show_error("Import Error", str(e))

    def display_contacts(self):
        self.tree.itemChanged.disconnect(self.handle_item_changed)
        self.tree.clear()
        
        for index, contact in enumerate(self.contacts, start=1):
            item = QTreeWidgetItem([
                str(index),
                contact.name,
                contact.phone or 'No Phone',
                contact.additional_phones or '-',
                '🖼️' if contact.has_photo else '',
            ])
            item.setCheckState(5, Qt.CheckState.Checked if contact.selected else Qt.CheckState.Unchecked)
            item.setData(0, Qt.ItemDataRole.UserRole, contact)
            
            # Set background color for selected contacts
            if contact.selected:
                selected_color = QColor(173, 216, 230)  # Light blue background
                for col in range(6):  # Apply to all columns
                    item.setBackground(col, QBrush(selected_color))
            else:
                # Reset to default background
                for col in range(6):
                    item.setBackground(col, QBrush())
            
            self.tree.addTopLevelItem(item)
        
        # Update header sort indicator
        if self.sort_column != 0:
            self.tree.header().setSortIndicator(self.sort_column, self.sort_order)
        else:
            self.tree.header().setSortIndicator(-1, self.sort_order)  # Clear sort indicator
        
        self.tree.itemChanged.connect(self.handle_item_changed)
        self.update_status_counts()

    def handle_item_changed(self, item, column):
        if column == 5:
            contact = item.data(0, Qt.ItemDataRole.UserRole)
            contact.selected = item.checkState(5) == Qt.CheckState.Checked
            
            # Update row background color based on selection
            if contact.selected:
                selected_color = QColor(173, 216, 230)  # Light blue background
                for col in range(6):  # Apply to all columns
                    item.setBackground(col, QBrush(selected_color))
            else:
                # Reset to default background
                for col in range(6):
                    item.setBackground(col, QBrush())
            
            self.update_status_counts()

    def sort_contacts(self, column):
        if column == self.sort_column:
            self.sort_order = (
                Qt.SortOrder.DescendingOrder 
                if self.sort_order == Qt.SortOrder.AscendingOrder 
                else Qt.SortOrder.AscendingOrder
            )
        else:
            self.sort_column = column
            self.sort_order = Qt.SortOrder.AscendingOrder

        if column == 0:  # Row number column - reset to original order
            self.contacts = self.all_contacts.copy()
            # Apply current search filter if active
            search_term = self.search_box.text().lower().replace('ي', 'ی').replace('ك', 'ک')
            if search_term:
                self.contacts = [
                    c for c in self.contacts
                    if (search_term in c.name.lower() or 
                        (c.phone and search_term in c.phone) or 
                        (c.additional_phones and search_term in c.additional_phones))
                ]
        elif column == 1:
            self.contacts.sort(key=lambda x: x.name.lower(), reverse=self.sort_order == Qt.SortOrder.DescendingOrder)
        elif column == 2:
            self.contacts.sort(key=lambda x: x.phone or '', reverse=self.sort_order == Qt.SortOrder.DescendingOrder)
        elif column == 3:
            self.contacts.sort(key=lambda x: len(x.additional_phones.split(', ')) if x.additional_phones else 0, reverse=self.sort_order == Qt.SortOrder.DescendingOrder)
        elif column == 4:
            self.contacts.sort(key=lambda x: not x.has_photo, reverse=self.sort_order == Qt.SortOrder.DescendingOrder)
        elif column == 5:
            self.contacts.sort(key=lambda x: not x.selected, reverse=self.sort_order == Qt.SortOrder.DescendingOrder)
            
        self.display_contacts()

    def filter_contacts(self):
        search_term = self.search_box.text().lower().replace('ي', 'ی').replace('ك', 'ک')
        if not search_term:
            self.contacts = self.all_contacts.copy()
        else:
            self.contacts = [
                c for c in self.all_contacts
                if (search_term in c.name.lower() or 
                    (c.phone and search_term in c.phone) or 
                    (c.additional_phones and search_term in c.additional_phones))
            ]
        self.display_contacts()

    def clear_search(self):
        self.search_box.clear()
        self.contacts = self.all_contacts.copy()
        self.display_contacts()

    def show_photo(self, item):
        contact = item.data(0, Qt.ItemDataRole.UserRole)
        if contact.photo_data:
            try:
                data = contact.photo_data
                missing_padding = len(data) % 4
                if missing_padding:
                    data += '=' * (4 - missing_padding)
                
                image_data = b64decode(data)
                
                header = image_data[:32].decode('ascii', errors='ignore')
                if 'JFIF' not in header and 'PNG' not in header:
                    raise ValueError("Invalid image format")
                
                image = Image.open(io.BytesIO(image_data))
                image.verify()
                image = Image.open(io.BytesIO(image_data))
                image.thumbnail((300, 300))
                
                qimage = QImage(image.tobytes(), image.width, image.height, QImage.Format.Format_RGB888)
                pixmap = QPixmap.fromImage(qimage)
                self.image_label.setPixmap(pixmap)
                self.status_bar.showMessage("Photo displayed")
            except Exception as e:
                self.show_error("Image Error", str(e))
        else:
            self.image_label.clear()
            self.status_bar.showMessage("No photo available")

    def delete_selected(self):
        selected_contacts = [c for c in self.all_contacts if c.selected]
        if not selected_contacts:
            self.show_warning("No Selection", "No contacts selected")
            return
            
        self.all_contacts = [c for c in self.all_contacts if not c.selected]
        self.contacts = self.all_contacts.copy()
        self.display_contacts()
        self.status_bar.showMessage(f"Deleted {len(selected_contacts)} contacts")

    def delete_contacts_without_phone(self):
        initial_count = len(self.contacts)
        self.contacts = [c for c in self.contacts if c.phone and c.phone.strip()]
        removed_count = initial_count - len(self.contacts)
        
        if removed_count > 0:
            self.all_contacts = [c for c in self.all_contacts if c.phone and c.phone.strip()]
            self.display_contacts()
            self.status_bar.showMessage(f"Removed {removed_count} contacts without phone numbers")
        else:
            self.status_bar.showMessage("No contacts without phone numbers found")

    def save_vcf(self):
        if not self.contacts:
            self.show_warning("Empty List", "No contacts to save")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save VCF", "", "VCF Files (*.vcf)")
        if not file_path:
            return

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                for contact in self.contacts:
                    in_photo = False
                    for line in contact.original_lines:
                        stripped_line = line.strip()
                        
                        if stripped_line == '':
                            f.write('\n')
                            continue
                            
                        if stripped_line.upper().startswith('PHOTO'):
                            f.write(stripped_line + '\n')
                            in_photo = True
                        elif in_photo:
                            if stripped_line.startswith('END:VCARD'):
                                f.write(line + '\n')
                                in_photo = False
                            else:
                                f.write(' ' + line.lstrip() + '\n')
                        else:
                            f.write(line + '\n')
            self.status_bar.showMessage("VCF saved successfully")
        except Exception as e:
            self.show_error("Saving Error", str(e))

    def show_error(self, title, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle(title)
        msg.setText(message)
        copy_btn = msg.addButton("Copy Error", QMessageBox.ButtonRole.ActionRole)
        msg.addButton(QMessageBox.StandardButton.Ok)
        msg.exec()

        if msg.clickedButton() == copy_btn:
            QGuiApplication.clipboard().setText(message)

    def show_warning(self, title, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle(title)
        msg.setText(message)
        copy_btn = msg.addButton("Copy Message", QMessageBox.ButtonRole.ActionRole)
        msg.addButton(QMessageBox.StandardButton.Ok)
        msg.exec()

        if msg.clickedButton() == copy_btn:
            QGuiApplication.clipboard().setText(message)

    def show_context_menu(self, position):
        item = self.tree.itemAt(position)
        if not item:
            return
            
        # Get the column that was clicked
        column = self.tree.columnAt(position.x())
        
        # Only show context menu for specific columns (0=# 1=Name, 2=Phone, 3=Additional Phones)
        if column not in [0, 1, 2, 3]:
            return
            
        # Get the text from the clicked cell
        cell_text = item.text(column)
        if not cell_text or cell_text in ['No Phone', '-']:
            return
            
        # Create context menu
        context_menu = QMenu(self)
        
        # Add copy action
        copy_text = f"Copy '{cell_text[:30]}{'...' if len(cell_text) > 30 else ''}'"
        copy_action = QAction(copy_text, self)
        copy_action.triggered.connect(lambda: self.copy_to_clipboard(cell_text))
        context_menu.addAction(copy_action)
        
        # Show the context menu
        context_menu.exec(self.tree.mapToGlobal(position))

    def copy_to_clipboard(self, text):
        QGuiApplication.clipboard().setText(text)
        self.status_bar.showMessage(f"Copied to clipboard: {text[:50]}{'...' if len(text) > 50 else ''}")

    def select_all(self):
        for contact in self.contacts:
            contact.selected = True
        self.display_contacts()
        self.status_bar.showMessage("All contacts selected")

    def deselect_all(self):
        for contact in self.contacts:
            contact.selected = False
        self.display_contacts()
        self.status_bar.showMessage("All contacts deselected")

    def invert_selection(self):
        for contact in self.contacts:
            contact.selected = not contact.selected
        self.display_contacts()
        self.status_bar.showMessage("Selection inverted")

    def update_status_counts(self):
        total_contacts = len(self.contacts)
        selected_contacts = len([c for c in self.contacts if c.selected])
        
        self.contact_count_label.setText(f"Contacts: {total_contacts}")
        self.selected_count_label.setText(f"Selected: {selected_contacts}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ContactViewer()
    window.show()
    sys.exit(app.exec())