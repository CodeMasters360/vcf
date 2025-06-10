import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from collections import Counter
import os
import traceback

class CopyableMessageBox:
    """A custom message box with copyable text"""
    
    def __init__(self, parent, title, message, msg_type="info"):
        self.result = None
        
        # Create top level window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x300")
        self.dialog.resizable(True, True)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center the dialog
        self.dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 50, parent.winfo_rooty() + 50))
        
        # Configure style based on message type
        if msg_type == "error":
            bg_color = "#ffebee"
            icon = "❌"
        elif msg_type == "warning":
            bg_color = "#fff3e0"
            icon = "⚠️"
        elif msg_type == "success":
            bg_color = "#e8f5e8"
            icon = "✅"
        else:
            bg_color = "#e3f2fd"
            icon = "ℹ️"
        
        self.dialog.configure(bg=bg_color)
        
        # Main frame
        main_frame = ttk.Frame(self.dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Icon and title frame
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(header_frame, text=icon, font=("Arial", 16)).pack(side=tk.LEFT)
        ttk.Label(header_frame, text=title, font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=(10, 0))
        
        # Message text (copyable)
        text_frame = ttk.LabelFrame(main_frame, text="Message (You can copy this text)", padding="10")
        text_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Text widget with scrollbar
        text_widget = tk.Text(text_frame, wrap=tk.WORD, height=8, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Insert message
        text_widget.insert(tk.END, message)
        text_widget.configure(state=tk.DISABLED)  # Make read-only but still copyable
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        # Copy button
        copy_btn = ttk.Button(button_frame, text="Copy to Clipboard", 
                             command=lambda: self.copy_to_clipboard(message))
        copy_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # OK button
        ok_btn = ttk.Button(button_frame, text="OK", command=self.ok_clicked)
        ok_btn.pack(side=tk.RIGHT)
        
        # Bind Enter key
        self.dialog.bind('<Return>', lambda e: self.ok_clicked())
        self.dialog.bind('<Escape>', lambda e: self.ok_clicked())
        
        # Focus on OK button
        ok_btn.focus()
        
        # Wait for dialog to close
        parent.wait_window(self.dialog)
    
    def copy_to_clipboard(self, text):
        """Copy text to clipboard"""
        self.dialog.clipboard_clear()
        self.dialog.clipboard_append(text)
        # Show brief confirmation
        original_text = self.dialog.title()
        self.dialog.title(f"{original_text} - Copied!")
        self.dialog.after(1000, lambda: self.dialog.title(original_text))
    
    def ok_clicked(self):
        """Handle OK button click"""
        self.result = True
        self.dialog.destroy()

class ExcelDuplicateFinder:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Duplicate Value Finder - یافتن مقادیر تکراری اکسل")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.file_path = tk.StringVar()
        self.duplicates_info = {}
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(main_frame, text="Excel Duplicate Value Finder", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="Select Excel File", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(file_frame, text="File Path:").grid(row=0, column=0, sticky=tk.W)
        
        path_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=60)
        path_entry.grid(row=0, column=1, padx=(10, 10), sticky=(tk.W, tk.E))
        
        browse_btn = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        browse_btn.grid(row=0, column=2)
        
        # Analysis button
        analyze_btn = ttk.Button(main_frame, text="Analyze for Duplicates", 
                                command=self.analyze_duplicates)
        analyze_btn.grid(row=2, column=0, columnspan=3, pady=10)
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Analysis Results", padding="10")
        results_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), 
                          pady=(10, 0))
        
        # Treeview for displaying results
        columns = ('Value', 'Count', 'Locations')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=15)
        
        # Define headings
        self.tree.heading('Value', text='Duplicate Value')
        self.tree.heading('Count', text='Count')
        self.tree.heading('Locations', text='Cell Locations')
        
        # Define column widths
        self.tree.column('Value', width=200)
        self.tree.column('Count', width=80)
        self.tree.column('Locations', width=400)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid the treeview and scrollbars
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Export button
        export_btn = ttk.Button(main_frame, text="Highlight & Save Excel", 
                               command=self.highlight_and_save)
        export_btn.grid(row=4, column=0, columnspan=3, pady=10)
        
        # Status label
        self.status_label = ttk.Label(main_frame, text="Ready to analyze Excel file")
        self.status_label.grid(row=5, column=0, columnspan=3, pady=(10, 0))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        file_frame.columnconfigure(1, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
    
    def show_error(self, title, message):
        """Show error message with copyable text"""
        CopyableMessageBox(self.root, title, message, "error")
    
    def show_warning(self, title, message):
        """Show warning message with copyable text"""
        CopyableMessageBox(self.root, title, message, "warning")
    
    def show_info(self, title, message):
        """Show info message with copyable text"""
        CopyableMessageBox(self.root, title, message, "info")
    
    def show_success(self, title, message):
        """Show success message with copyable text"""
        CopyableMessageBox(self.root, title, message, "success")
    
    def browse_file(self):
        """Open file dialog to select Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path.set(file_path)
    
    def analyze_duplicates(self):
        """Analyze the Excel file for duplicate values"""
        if not self.file_path.get():
            self.show_error("Error", "Please select an Excel file first!")
            return
        
        try:
            self.status_label.config(text="Analyzing file...")
            self.root.update()
            
            # Clear previous results
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Read all sheets from Excel file
            excel_file = pd.ExcelFile(self.file_path.get())
            all_values = {}  # Dictionary to store value: [(sheet, row, col), ...]
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(self.file_path.get(), sheet_name=sheet_name, header=None)
                
                # Iterate through all cells
                for row_idx in range(len(df)):
                    for col_idx in range(len(df.columns)):
                        cell_value = df.iloc[row_idx, col_idx]
                        
                        # Skip NaN values
                        if pd.isna(cell_value):
                            continue
                        
                        # Convert to string for consistency
                        cell_value = str(cell_value)
                        
                        # Skip empty strings
                        if cell_value.strip() == '':
                            continue
                        
                        # Store location information
                        location = (sheet_name, row_idx + 1, col_idx + 1)  # +1 for 1-based indexing
                        
                        if cell_value not in all_values:
                            all_values[cell_value] = []
                        all_values[cell_value].append(location)
            
            # Find duplicates (values that appear more than once)
            duplicates = {value: locations for value, locations in all_values.items() 
                         if len(locations) > 1}
            
            self.duplicates_info = duplicates
            
            # Populate the treeview
            for value, locations in duplicates.items():
                # Format locations as readable string
                location_strs = []
                for sheet, row, col in locations:
                    col_letter = self.number_to_column_letter(col)
                    location_strs.append(f"{sheet}!{col_letter}{row}")
                
                locations_text = ", ".join(location_strs)
                
                self.tree.insert('', tk.END, values=(value, len(locations), locations_text))
            
            # Update status
            duplicate_count = len(duplicates)
            total_duplicate_cells = sum(len(locations) for locations in duplicates.values())
            
            self.status_label.config(
                text=f"Found {duplicate_count} duplicate values in {total_duplicate_cells} cells"
            )
            
            if duplicate_count == 0:
                self.show_info("Results", "No duplicate values found in the Excel file!")
            
        except Exception as e:
            error_details = f"Error details:\n{str(e)}\n\nFull traceback:\n{traceback.format_exc()}"
            self.show_error("Analysis Error", 
                          f"An error occurred while analyzing the file:\n\n{error_details}")
            self.status_label.config(text="Error occurred during analysis")
    
    def number_to_column_letter(self, col_num):
        """Convert column number to Excel column letter (e.g., 1 -> A, 27 -> AA)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def highlight_and_save(self):
        """Highlight duplicate cells in Excel and save to a new file"""
        if not self.duplicates_info:
            self.show_warning("Warning", "No duplicates to highlight! Please analyze first.")
            return
        
        try:
            self.status_label.config(text="Highlighting duplicates and saving...")
            self.root.update()
            
            # Load the workbook
            workbook = openpyxl.load_workbook(self.file_path.get())
            
            # Define highlight color (yellow)
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Highlight all duplicate cells
            for value, locations in self.duplicates_info.items():
                for sheet_name, row, col in locations:
                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = highlight_fill
            
            # Save to a new file
            original_path = self.file_path.get()
            name, ext = os.path.splitext(original_path)
            new_path = f"{name}_duplicates_highlighted{ext}"
            
            workbook.save(new_path)
            
            self.status_label.config(text=f"Highlighted file saved as: {os.path.basename(new_path)}")
            self.show_success("Success", 
                            f"Duplicate values have been highlighted and saved to:\n\n{new_path}\n\nAll duplicate cells are now highlighted in yellow color.")
            
        except Exception as e:
            error_details = f"Error details:\n{str(e)}\n\nFull traceback:\n{traceback.format_exc()}"
            self.show_error("Highlighting Error", 
                          f"An error occurred while highlighting:\n\n{error_details}")
            self.status_label.config(text="Error occurred during highlighting")

def main():
    root = tk.Tk()
    app = ExcelDuplicateFinder(root)
    root.mainloop()

if __name__ == "__main__":
    main()